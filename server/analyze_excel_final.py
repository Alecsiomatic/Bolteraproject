import openpyxl
from collections import defaultdict

print("=" * 100)
print("ANÁLISIS DETALLADO DE NUMERACIÓN - HOJA FINAL DEL EXCEL")
print("=" * 100)

# Buscar el archivo Excel
import os
excel_path = None
for root, dirs, files in os.walk(r"C:\Users\Alecs\Desktop\ddu\BOLETERA PROJECT"):
    for f in files:
        if "ZONIFICACION" in f.upper() and f.endswith(('.xlsx', '.xls')):
            excel_path = os.path.join(root, f)
            print(f"Encontrado: {excel_path}")

if not excel_path:
    # Usar el PDF entonces
    print("No se encontró Excel, analizando desde el PDF...")
    exit()

wb = openpyxl.load_workbook(excel_path)
print(f"\nHojas disponibles: {wb.sheetnames}")

# Buscar hoja FINAL
sheet = None
for name in wb.sheetnames:
    if 'FINAL' in name.upper():
        sheet = wb[name]
        print(f"\n✅ Usando hoja: {name}")
        break

if not sheet:
    print("No se encontró hoja FINAL")
    sheet = wb.active
    print(f"Usando hoja activa: {sheet.title}")

# Leer datos
print("\n" + "=" * 100)
print("📊 ESTRUCTURA DE DATOS EN LA HOJA FINAL")
print("=" * 100)

# Imprimir las primeras filas para entender la estructura
print("\nPrimeras 30 filas y columnas A-J:")
for row in range(1, 31):
    row_data = []
    for col in range(1, 11):  # A-J
        cell = sheet.cell(row=row, column=col)
        val = cell.value
        if val is not None:
            row_data.append(f"{openpyxl.utils.get_column_letter(col)}:{str(val)[:15]}")
    if row_data:
        print(f"  Fila {row:2d}: {', '.join(row_data)}")

# Buscar secciones específicas
print("\n" + "=" * 100)
print("🔍 BUSCANDO INFORMACIÓN DE SECCIONES Y NUMERACIÓN")
print("=" * 100)

sections_data = {}
current_section = None
current_zone = None

# Recorrer todo el Excel buscando patrones
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        val = str(cell.value).upper() if cell.value else ""
        
        # Detectar zona
        if "VIP" in val and len(val) < 20:
            current_zone = "VIP"
        elif "PLUS" in val and len(val) < 20:
            current_zone = "PLUS"
        elif "PREFERENTE" in val and len(val) < 20:
            current_zone = "PREFERENTE"
        
        # Detectar sección
        if "IZQUIERDA" in val or "IZQUIER" in val:
            current_section = "IZQUIERDA"
        elif "CENTRAL" in val:
            current_section = "CENTRAL"
        elif "DERECHA" in val:
            current_section = "DERECHA"

# Buscar específicamente rangos de asientos
print("\n🔎 Buscando patrones de numeración (rangos tipo '1-43', 'Del 1 al 43', etc.):")

for row in range(1, min(100, sheet.max_row + 1)):
    for col in range(1, min(20, sheet.max_column + 1)):
        cell = sheet.cell(row=row, column=col)
        val = str(cell.value) if cell.value else ""
        
        # Buscar patrones de rangos
        if '-' in val and any(c.isdigit() for c in val):
            if len(val) < 30:
                print(f"  [{row},{col}] {val}")
        elif 'del' in val.lower() and 'al' in val.lower():
            print(f"  [{row},{col}] {val}")

# Buscar la tabla de asientos por sección
print("\n" + "=" * 100)
print("📋 ESTRUCTURA DE ASIENTOS POR SECCIÓN (según análisis previo)")
print("=" * 100)

# Datos conocidos del análisis previo del PDF
teatro_structure = {
    "VIP": {
        "filas": ["8", "7", "6", "5", "4", "3", "2", "1"],  # 8 filas
        "asientos_por_seccion": {
            "DERECHA": 182,    # X bajo en el canvas
            "CENTRAL": 144,
            "IZQUIERDA": 188,  # X alto en el canvas
        },
        "total": 514
    },
    "PLUS": {
        "filas": ["P", "O", "N", "M", "L", "K", "J", "I", "H", "G", "F", "E", "D", "C", "B", "A"],  # 16 filas
        "asientos_por_seccion": {
            "DERECHA": 435,
            "CENTRAL": 414,
            "IZQUIERDA": 445,
        },
        "total": 1294
    },
    "PREFERENTE": {
        "filas": ["P", "O", "N", "M", "L", "K", "J", "I", "H", "G", "F", "E", "D", "C", "B", "A"],  # 16 filas
        "asientos_por_seccion": {
            "DERECHA": 631,
            "CENTRAL": 792,
            "IZQUIERDA": 638,
        },
        "total": 2061
    }
}

print("\n🎭 DISTRIBUCIÓN DE ASIENTOS POR ZONA Y SECCIÓN:")
print("-" * 80)
print(f"{'ZONA':<15} | {'DERECHA':<12} | {'CENTRAL':<12} | {'IZQUIERDA':<12} | {'TOTAL':<10}")
print("-" * 80)

total_derecha = 0
total_central = 0
total_izquierda = 0
total_total = 0

for zona, data in teatro_structure.items():
    d = data["asientos_por_seccion"]["DERECHA"]
    c = data["asientos_por_seccion"]["CENTRAL"]
    i = data["asientos_por_seccion"]["IZQUIERDA"]
    t = data["total"]
    print(f"{zona:<15} | {d:<12} | {c:<12} | {i:<12} | {t:<10}")
    total_derecha += d
    total_central += c
    total_izquierda += i
    total_total += t

print("-" * 80)
print(f"{'TOTAL':<15} | {total_derecha:<12} | {total_central:<12} | {total_izquierda:<12} | {total_total:<10}")

print("\n" + "=" * 100)
print("🔢 ANÁLISIS DE GAPS EN LA NUMERACIÓN POR FILA")
print("=" * 100)

print("""
La numeración de asientos es CONTINUA a través de todas las secciones de cada fila.
Es decir, si una fila tiene 143 asientos:

  ┌──────────────┬───────────────────────┬──────────────┐
  │   DERECHA    │       CENTRAL         │  IZQUIERDA   │
  │   (1-43)     │      (44-100)         │  (101-143)   │
  └──────────────┴───────────────────────┴──────────────┘
  
        ↓                  ↓                    ↓
    Números bajos    Números medios      Números altos
    
  * Nota: DERECHA del espectador = X BAJO en el canvas (lado izquierdo visual)
  * IZQUIERDA del espectador = X ALTO en el canvas (lado derecho visual)
""")

# Calcular gaps por fila basado en los totales
print("\n📊 CÁLCULO ESTIMADO DE GAPS POR FILA:")
print("=" * 80)

# VIP - 8 filas, 514 asientos total
vip_promedio_fila = 514 // 8  # ~64 asientos por fila
print(f"\n🔵 VIP (8 filas, ~{vip_promedio_fila} asientos/fila):")
print(f"   DERECHA:   asientos {1}-{182//8} (promedio por fila)")
print(f"   CENTRAL:   asientos {182//8+1}-{(182+144)//8}")
print(f"   IZQUIERDA: asientos {(182+144)//8+1}-{vip_promedio_fila}")

# PLUS - 16 filas, 1294 asientos total
plus_promedio_fila = 1294 // 16  # ~81 asientos por fila
print(f"\n🟢 PLUS (16 filas, ~{plus_promedio_fila} asientos/fila):")
print(f"   DERECHA:   asientos del rango bajo")
print(f"   CENTRAL:   asientos del rango medio")
print(f"   IZQUIERDA: asientos del rango alto")

# PREFERENTE - 16 filas, 2061 asientos total
pref_promedio_fila = 2061 // 16  # ~129 asientos por fila
print(f"\n🟡 PREFERENTE (16 filas, ~{pref_promedio_fila} asientos/fila):")
print(f"   DERECHA:   asientos del rango bajo")
print(f"   CENTRAL:   asientos del rango medio")
print(f"   IZQUIERDA: asientos del rango alto")

print("\n" + "=" * 100)
print("⚠️  NOTA IMPORTANTE SOBRE GAPS:")
print("=" * 100)
print("""
Para determinar los GAPS exactos (dónde empieza y termina cada sección en cada fila),
necesito ver los datos específicos del Excel hoja FINAL.

Lo que puedo confirmar de la estructura:

1. NUMERACIÓN CONTINUA: Los asientos se numeran de forma continua en cada fila
   (no reinician a 1 en cada sección)

2. DIRECCIÓN: 
   - Números BAJOS = lado DERECHA del espectador (mirando al escenario)
   - Números ALTOS = lado IZQUIERDA del espectador

3. GAPS ENTRE SECCIONES: Los "gaps" en la numeración indican dónde termina
   una sección y empieza otra. Por ejemplo:
   
   Si DERECHA tiene asientos 1-40 y CENTRAL tiene 45-90:
   - Gap de 41-44 = pasillo entre DERECHA y CENTRAL
   
4. FILAS INCOMPLETAS: Algunas filas pueden tener menos asientos debido a:
   - Pasillos
   - Forma del auditorio (filas laterales más cortas)
   - Zonas técnicas o de acceso
""")

wb.close()
print("\n✅ Análisis completado")
