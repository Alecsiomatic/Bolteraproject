#!/usr/bin/env python3
"""
Leer Excel de asientos y mostrar la estructura de numeración
"""
import subprocess
import sys

# Instalar openpyxl si no existe
subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])

from openpyxl import load_workbook
import json

wb = load_workbook('/tmp/tangamanga.xlsx', data_only=True)

print("Hojas disponibles:", wb.sheetnames)
print()

all_sections = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== {sheet_name} ===")
    
    # Leer todas las filas
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if row[0] is not None:  # Si hay dato en la primera columna
            fila = row[0]
            asientos = row[1] if len(row) > 1 and row[1] else 0
            
            # Buscar números de butaca específicos en las siguientes columnas
            seat_numbers = []
            if len(row) > 2:
                for i in range(2, len(row)):
                    if row[i] is not None:
                        try:
                            seat_numbers.append(int(row[i]))
                        except (ValueError, TypeError):
                            pass
            
            if asientos and int(asientos) > 0:
                rows_data.append({
                    'fila': fila,
                    'asientos': int(asientos),
                    'seat_numbers': seat_numbers if seat_numbers else list(range(1, int(asientos) + 1))
                })
                
                nums = seat_numbers if seat_numbers else list(range(1, int(asientos) + 1))
                if len(nums) > 10:
                    print(f"  Fila {fila}: {asientos} asientos -> [{nums[0]}, {nums[1]}, ... {nums[-2]}, {nums[-1]}]")
                else:
                    print(f"  Fila {fila}: {asientos} asientos -> {nums}")
    
    all_sections[sheet_name] = {'filas': rows_data}
    print()

# Guardar JSON
with open('/tmp/tangamanga_seats.json', 'w') as f:
    json.dump(all_sections, f, indent=2, ensure_ascii=False)

print("JSON guardado en /tmp/tangamanga_seats.json")
