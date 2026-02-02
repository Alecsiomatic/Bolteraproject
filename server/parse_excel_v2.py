#!/usr/bin/env python3
"""
Leer Excel con la estructura correcta - CORREGIDO
Usa la cantidad de asientos como autoridad, no el rango de numeración
"""
from openpyxl import load_workbook
import json
import re

wb = load_workbook('/tmp/tangamanga.xlsx', data_only=True)
ws = wb['VERTICAL']

all_sections = {}
current_section = None
current_section_data = {'filas': [], 'total': 0}

def parse_row_data(row):
    """Parsea los datos de una fila del Excel"""
    fila = row[3]
    num_asientos = int(row[4]) if row[4] else 0
    direccion = str(row[5]) if row[5] else 'IZQ A DERECHA'
    numeracion = str(row[6]) if row[6] else ''
    
    # Parsear numeración (ej: "1 a 22", "23 a 37", "38 a 59")
    match = re.search(r'(\d+)\s*a\s*(\d+)', numeracion)
    if match:
        start = int(match.group(1))
        end = int(match.group(2))
        
        # Generar exactamente num_asientos números
        # Si start < end, van en orden ascendente; si start > end, descendente
        if start <= end:
            # Ascendente: 1 a 22 -> [1,2,3,...,22]
            seat_numbers = list(range(start, start + num_asientos))
        else:
            # Descendente: 37 a 23 -> empezar en 37 e ir bajando
            seat_numbers = list(range(start, start - num_asientos, -1))
    else:
        # Sin numeración especificada, usar 1 a N
        seat_numbers = list(range(1, num_asientos + 1))
    
    # Si dirección es DERECHA A IZQ, los asientos se leen de derecha a izquierda
    # Los números ya están en el orden físico correcto, solo invertimos el orden de renderizado
    if 'DERECHA' in direccion.upper() and 'IZQ' in direccion.upper():
        seat_numbers = seat_numbers[::-1]  # Invertir para que el primer asiento esté a la derecha
    
    return {
        'fila': fila,
        'asientos': num_asientos,
        'direccion': direccion,
        'seat_numbers': seat_numbers
    }

for row in ws.iter_rows(min_row=2, values_only=True):
    # Buscar inicio de sección (tiene nombre en columna B)
    if row[1] and 'SECC.' in str(row[1]):
        # Guardar sección anterior
        if current_section and current_section_data['filas']:
            all_sections[current_section] = current_section_data

        # Nueva sección
        section_name = str(row[1]).replace('SECC. ', '').replace('SECC.', '').strip()
        current_section = section_name
        total = row[2] if row[2] else 0
        current_section_data = {'filas': [], 'total': int(total) if total else 0}

        # Esta fila también tiene datos de la primera fila
        if row[3] is not None:
            row_data = parse_row_data(row)
            current_section_data['filas'].append(row_data)

    # Fila de datos (sin nombre de sección)
    elif current_section and row[3] is not None:
        num_asientos = int(row[4]) if row[4] else 0
        if num_asientos > 0:
            row_data = parse_row_data(row)
            current_section_data['filas'].append(row_data)

# Guardar última sección
if current_section and current_section_data['filas']:
    all_sections[current_section] = current_section_data

# Mostrar resumen con verificación
print("=== RESUMEN DE SECCIONES (CORREGIDO) ===\n")
total_seats = 0
for section, data in all_sections.items():
    calc_total = sum(f['asientos'] for f in data['filas'])
    total_seats += calc_total
    status = "✓" if calc_total == data['total'] else f"✗ (calc={calc_total})"
    print(f"{section}: {data['total']} declarado {status}")
    for fila in data['filas']:
        nums = fila['seat_numbers']
        nums_ok = "✓" if len(nums) == fila['asientos'] else "✗"
        print(f"  Fila {fila['fila']}: {fila['asientos']} asientos {nums_ok}, nums: {nums[0]} → {nums[-1]} ({fila['direccion']})")
    print()

print(f"TOTAL GENERAL: {total_seats} asientos")

# Guardar JSON
with open('/tmp/tangamanga_seats.json', 'w') as f:
    json.dump(all_sections, f, indent=2, ensure_ascii=False)

print("\nJSON guardado en /tmp/tangamanga_seats.json")
