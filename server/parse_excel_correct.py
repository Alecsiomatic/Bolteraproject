#!/usr/bin/env python3
"""
Leer Excel con la estructura correcta de numeración
"""
from openpyxl import load_workbook
import json
import re

wb = load_workbook('/tmp/tangamanga.xlsx', data_only=True)

# Usar la hoja VERTICAL que tiene mejor estructura
ws = wb['VERTICAL']

all_sections = {}
current_section = None
current_section_data = {'filas': [], 'total': 0}

for row in ws.iter_rows(min_row=2, values_only=True):
    # Buscar inicio de sección (tiene nombre en columna B)
    if row[1] and 'SECC.' in str(row[1]):
        # Guardar sección anterior
        if current_section and current_section_data['filas']:
            all_sections[current_section] = current_section_data
        
        # Nueva sección
        section_name = str(row[1]).replace('SECC. ', '').replace('SECC.', '').strip()
        current_section = section_name
        total = row[2] if row[2] else 0
        current_section_data = {'filas': [], 'total': int(total) if total else 0}
        
        # Esta fila también tiene datos de la primera fila
        if row[3] is not None:  # Columna FILA
            fila = row[3]
            num_asientos = int(row[4]) if row[4] else 0
            direccion = str(row[5]) if row[5] else 'IZQ A DERECHA'
            numeracion = str(row[6]) if row[6] else ''
            
            # Parsear numeración (ej: "1 a 22", "38 a 59")
            match = re.search(r'(\d+)\s*a\s*(\d+)', numeracion)
            if match:
                start = int(match.group(1))
                end = int(match.group(2))
                seat_numbers = list(range(start, end + 1))
            else:
                seat_numbers = list(range(1, num_asientos + 1))
            
            # Si dirección es DERECHA A IZQ, los números van de mayor a menor
            if 'DERECHA' in direccion.upper() and 'IZQ' in direccion.upper():
                seat_numbers = seat_numbers[::-1]  # Invertir
            
            current_section_data['filas'].append({
                'fila': fila,
                'asientos': num_asientos,
                'direccion': direccion,
                'seat_numbers': seat_numbers
            })
    
    # Fila de datos (sin nombre de sección)
    elif current_section and row[3] is not None:
        fila = row[3]
        num_asientos = int(row[4]) if row[4] else 0
        direccion = str(row[5]) if row[5] else 'IZQ A DERECHA'
        numeracion = str(row[6]) if row[6] else ''
        
        # Parsear numeración
        match = re.search(r'(\d+)\s*a\s*(\d+)', numeracion.replace('a', ' a '))
        if match:
            start = int(match.group(1))
            end = int(match.group(2))
            seat_numbers = list(range(start, end + 1))
        else:
            seat_numbers = list(range(1, num_asientos + 1))
        
        # Si dirección es DERECHA A IZQ, los números van de mayor a menor
        if 'DERECHA' in direccion.upper() and 'IZQ' in direccion.upper():
            seat_numbers = seat_numbers[::-1]  # Invertir
        
        if num_asientos > 0:
            current_section_data['filas'].append({
                'fila': fila,
                'asientos': num_asientos,
                'direccion': direccion,
                'seat_numbers': seat_numbers
            })

# Guardar última sección
if current_section and current_section_data['filas']:
    all_sections[current_section] = current_section_data

# Mostrar resumen
print("=== RESUMEN DE SECCIONES ===\n")
for section, data in all_sections.items():
    print(f"{section}: {data['total']} total, {len(data['filas'])} filas")
    for fila in data['filas'][:2]:  # Primeras 2 filas como ejemplo
        nums = fila['seat_numbers']
        print(f"  Fila {fila['fila']}: {fila['asientos']} asientos, nums: {nums[0]} a {nums[-1]} ({fila['direccion']})")
    if len(data['filas']) > 2:
        print(f"  ... y {len(data['filas']) - 2} filas más")
    print()

# Guardar JSON
with open('/tmp/tangamanga_seats.json', 'w') as f:
    json.dump(all_sections, f, indent=2, ensure_ascii=False)

print("JSON guardado en /tmp/tangamanga_seats.json")
