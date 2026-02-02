#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('/tmp/tangamanga.xlsx', data_only=True)

print("Hojas:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== {sheet_name} ===")
    print(f"Dimensiones: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Leer primeras 30 filas
    print("Contenido:")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True)):
        if any(cell is not None for cell in row):
            print(f"  Fila {i+1}: {row[:10]}")  # Primeras 10 columnas
    print()
