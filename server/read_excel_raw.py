#!/usr/bin/env python3
"""
Leer Excel crudo para ver exactamente qué dice
"""
from openpyxl import load_workbook

wb = load_workbook('/tmp/tangamanga.xlsx', data_only=True)
ws = wb['VERTICAL']

print("=== CONTENIDO CRUDO DE LA HOJA VERTICAL ===\n")
print("Columnas: A=1, B=ZONA, C=TOTAL, D=FILA, E=NO.ASIENTOS, F=DIRECCION, G=NUMERACION")
print("-" * 100)

for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    # Solo mostrar filas con datos relevantes
    if row[1] or row[3]:  # Si tiene zona o fila
        print(f"Fila {row_num:2d}: Zona={row[1]!r:25} Total={row[2]!r:5} Fila={row[3]!r:5} Asientos={row[4]!r:5} Dir={row[5]!r:20} Num={row[6]!r}")
